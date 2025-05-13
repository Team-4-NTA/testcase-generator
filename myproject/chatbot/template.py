import os
import json
import re
import openai
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from shutil import copyfile
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from openpyxl.styles import Alignment

from .models import Chat, ChatDetail

client = openai.OpenAI(api_key="")

# Đường dẫn thư mục lưu file
MEDIA_DIR = Path("media")
MEDIA_DIR.mkdir(exist_ok=True)


@csrf_exempt
def generate_template(request):
    data = json.loads(request.body)
    screen_name = data.get("screen_name")
    requirement = data.get("requirement")
    type = data.get("type")
    history_id = data.get('history_id')
    if type == "spec":
        template_data = generate_spec_data(data)
        file_path, file_name = create_excel_file_spec(screen_name, requirement, template_data)
    elif type == "api":
        template_data, data_response = generate_api_data(data)
        file_path, file_name = create_excel_file_api(screen_name, template_data)
    else:
        return JsonResponse({
            "error": "Invalid type provided. Valid types are 'spec' or 'api'."
        }, status=400)
    file_path = str(file_path).replace(str(settings.BASE_DIR), "").lstrip("/")
    save_template(file_path, file_name, screen_name, requirement, history_id)
    return JsonResponse({
        "screen_name": screen_name,
        "file_name": file_name,
        "data": template_data,
        "file_url": str(file_path)
    })

def generate_spec_data(data):
    screen_name = data.get('screen_name', '')
    requirement = data.get('requirement', '')
    prompt = (
        f"Dựa vào màn hình {screen_name} và yêu cầu: {requirement}, tạo danh sách thông số cần có:\n"
        "- STT\n- Tên Item\n- Require\n- Type\n- Max\n- Min\n- Condition/Spec/Event\n- Data"
    )

    response = client.chat.completions.create(
        model="gpt-4o-mini-2024-07-18",
        messages=[
            {"role": "user", "content": prompt}
        ],
    )

    return parse_spec_data(response.choices[0].message.content)

def parse_spec_data(response_text):
    """Chuyển đổi bảng markdown từ OpenAI thành danh sách dictionary"""
    lines = response_text.split("\n")

    spec_data = []
    table_started = False

    for line in lines:
        # Bỏ qua tiêu đề và đường phân cách bảng
        if "| STT |" in line:
            table_started = True
            continue
        if table_started and re.match(r"^\|\s*-+\s*\|", line):
            continue

        # Nếu bảng đã bắt đầu, xử lý từng dòng
        if table_started and "|" in line:
            columns = [col.strip() for col in line.split("|")[1:-1]]  # Bỏ cột đầu và cuối (do split("|") tạo phần tử rỗng)
            if len(columns) == 8:  # Đúng số lượng cột
                spec_data.append({
                    "STT": columns[0],
                    "Tên Item": columns[1],
                    "Require": columns[2],
                    "Type": columns[3],
                    "Max": columns[4],
                    "Min": columns[5],
                    "Condition/Spec/Event": columns[6],
                    "Data": columns[7]
                })

    return spec_data

def copy_file(type = "spec"):
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{current_time}_{type}.xlsx"
    static_dir = os.path.join(settings.BASE_DIR, 'static')
    outputs_dir = os.path.join(static_dir, 'outputs')
    os.makedirs(outputs_dir, exist_ok=True)
    
    template_path = os.path.join(static_dir, f'{type}.xlsx')

    if not os.path.exists(template_path):
        return JsonResponse({"error": f"Không tìm thấy file mẫu {type}.xlsx"}, status=500)
    
    output_path = os.path.join(outputs_dir, file_name)
    
    try:
        copyfile(template_path, output_path)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"Template file not found at {template_path}. "
            "Please ensure the file exists in static directory"
        )
    
    return file_name, output_path

def create_excel_file_spec(screen_name: str, requirement: str, spec_data: list, creator: str = "System", version: str = "1"):
    file_name, output_path = copy_file()

    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # Write metadata
        ws["B1"] = screen_name
        ws["B2"] = requirement
        ws["B3"] = creator
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["B5"] = version

        start_row = 9
        for i, item in enumerate(spec_data):
            row = [
                item.get("STT", ""),
                item.get("Tên Item", ""),
                item.get("Require", ""),
                item.get("Type", ""),
                item.get("Max", ""),
                item.get("Min", ""),
                item.get("Condition/Spec/Event", ""),
                item.get("Data", "")
            ]
            for col, value in enumerate(row, start=1):
                ws.cell(row=start_row + i, column=col, value=value)

        wb.save(output_path)
        
        return f"/static/outputs/{file_name}", file_name

    except Exception as e:
        return {"error": f"Error creating Excel file: {str(e)}"}

def generate_api_data(data):
    screen_name = data.get('screen_name', '')
    requirement = data.get('requirement', '')
    prompt = (
        f"Hãy thiết kế tài liệu mô tả cho API với các thông tin sau:\n\n"
        f"- Tên API: {screen_name}\n"
        f"- Yêu cầu chức năng: {requirement}\n\n"
        
        "Trả về nội dung theo cấu trúc gồm hai phần:\n"
        
        "1. Thông tin tổng quan (meta), trình bày theo định dạng:\n"
        "Tên API: <tên API>\n"
        "Phương thức yêu cầu: <GET/POST/...>\n"
        "URL yêu cầu: <đường dẫn endpoint>\n"
        "Giao thức truyền thông: HTTP hoặc HTTPS\n"
        "Phương thức phản hồi: JSON hoặc XML\n\n"
        
        "2. Danh sách tham số, chia làm hai nhóm rõ ràng:\n"
        
        "### Request Parameters\n"
        "Trình bày từng tham số trên một dòng, với các cột sau, cột Chi tiết hãy cho biết cả validation:\n"
        "- STT\n- Tên tham số\n- Chi tiết (kiểu dữ liệu hoặc mô tả cấu trúc)\n- Bắt buộc (có hoặc không)\n- Mô tả\n- Mẫu dữ liệu\n\n"
        "Ví dụ:\n"
        "| 1 | user_id | integer, không thể null, thuộc id của table users | có | ID người dùng | 123 |\n"
        "| 2 | token | string, không thể null | có | Mã xác thực người dùng | abc123 |\n\n"
        
        "### Response Parameters\n"
        "Trình bày tương tự như bảng request, ví dụ:\n"
        "| STT | Tên tham số | Chi tiết | Bắt buộc | Mô tả | Mẫu dữ liệu |\n"
        "|-----|-------------|----------|----------|--------|--------------|\n"
        "| 1 | status | string | có | Trạng thái kết quả | success |\n"
        "| 2 | data | object | có | Thông tin chi tiết người dùng | { \"user_id\": 1, \"username\": \"user123\" } |\n\n"
    )

    response = client.chat.completions.create(
        model="gpt-4o-mini-2024-07-18",
        messages=[
            {"role": "user", "content": prompt}
        ],
    )

    return parse_api_data(response.choices[0].message.content)

def parse_api_data(response_text):
    lines = [line.strip() for line in response_text.strip().split("\n") if line.strip()]
    
    meta = {}
    request_params = []
    response_params = []

    current_section = None
    in_table = False
    table_header = None

    for line in lines:
        if line == "## Thông tin tổng quan (meta)":
            current_section = "meta"
            continue
        elif "Request Parameters" in line:
            current_section = "request"
            continue
        elif "Response Parameters" in line:
            current_section = "response"
            continue

        if current_section == "meta" and line.startswith("- "):
            parts = [p.strip() for p in line[2:].split(":", 1)]
            if len(parts) == 2:
                key, value = parts
                meta[key] = value

        if current_section in ["request", "response"]:
            if line.startswith("|") and "STT" in line and "Tên tham số" in line:
                in_table = True
                table_header = [h.strip() for h in line.strip('|').split('|')]
                continue
            
            if in_table and not line.startswith("|"):
                in_table = False
                continue
            
            if in_table and line.startswith("|") and not line.startswith("|---"):
                columns = [col.strip() for col in line.strip('|').split('|')]
                if len(columns) == len(table_header):
                    param = dict(zip(table_header, columns))
                    if current_section == "request":
                        request_params.append(param)
                    elif current_section == "response":
                        response_params.append(param)

    return {
        "meta": meta,
        "request_params": request_params,
        "response_params": response_params
    }

def copy_row(ws, source_row: int, target_row: int):
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)

        # Handle merge of B (col=2) and C (col=3)
        if col == 2:
            source_cell_b = ws.cell(row=source_row, column=2)
            source_cell_c = ws.cell(row=source_row, column=3)
            
            # Combine values
            combined_value = f"{source_cell_b.value or ''} {source_cell_c.value or ''}".strip()
            
            target_cell = ws.cell(row=target_row, column=2)
            target_cell.value = combined_value

            # Copy style from column B (optional: could be improved)
            if source_cell_b.has_style:
                target_cell.font = copy(source_cell_b.font)
                target_cell.border = copy(source_cell_b.border)
                target_cell.fill = copy(source_cell_b.fill)
                target_cell.number_format = copy(source_cell_b.number_format)
                target_cell.protection = copy(source_cell_b.protection)
                target_cell.alignment = copy(source_cell_b.alignment)

            # Merge B and C in target row
            ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=3)

        elif col == 3:
            continue  # Skip column C since it's merged with B

        else:
            target_cell = ws.cell(row=target_row, column=col)
            target_cell.value = source_cell.value

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

def write_params(sheet, params: list, start_row: int) -> int:
    for i, param in enumerate(params, start=start_row):
        sheet.cell(row=i, column=2).value = param.get("Tên tham số")
        sheet.cell(row=i, column=3).value = param.get("Mô tả")
        sheet.cell(row=i, column=4).value = param.get("Bắt buộc")
        sheet.cell(row=i, column=5).value = param.get("Chi tiết")
        sheet.cell(row=i, column=6).value = param.get("Mẫu dữ liệu")
    return start_row + len(params) - 1 

def create_excel_file_api(screen_name: str, api_data: dict):
    file_name, output_path = copy_file("api")
    
    try:
        wb = load_workbook(output_path)
        sheet = wb.active
        
        # Ghi meta data vào các ô cụ thể
        meta = api_data.get("meta", {})
        meta_fields = {
            "B1": ("Tên API", screen_name),
            "B3": ("Phương thức yêu cầu", "GET"),
            "B4": ("URL yêu cầu", "/your/api/url"),
            "B5": ("Giao thức truyền thông", "HTTPS"),
            "B6": ("Phương thức phản hồi", "JSON")
        }
        
        for cell, (field, default) in meta_fields.items():
            sheet[cell] = meta.get(field) or default
        
        # Ghi request parameters
        request_params = api_data.get("request_params", [])
        start_row = 10
        current_row = write_params(sheet, request_params, start_row=start_row)
        
        # Merge ô cột A cho phần Request Parameters, căn giữa
        if current_row > start_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=current_row, end_column=1)
        cell_request = sheet.cell(row=start_row, column=1)
        cell_request.value = "Request"
        cell_request.alignment = Alignment(horizontal='center', vertical='center')
        
        # Copy header (dòng 9) xuống dưới phần Request (ở dòng current_row + 2)
        copy_row(sheet, source_row=9, target_row=current_row + 2)
        
        # Ghi response parameters
        response_start_row = current_row + 3
        response_params = api_data.get("response_params", [])
        response_end_row = write_params(sheet, response_params, start_row=response_start_row)
        
        if response_end_row > response_start_row:
            sheet.merge_cells(start_row=response_start_row, start_column=1, end_row=response_end_row, end_column=1)
        cell_response = sheet.cell(row=response_start_row, column=1)
        cell_response.value = "Response"
        cell_response.alignment = Alignment(horizontal='center', vertical='center')

        
        wb.save(output_path)
        return output_path, file_name
    
    except Exception as e:
        if os.path.exists(output_path):
            os.remove(output_path)
        raise RuntimeError(f"Failed to process Excel file: {str(e)}")

def save_template(file_path, file_name, screen_name, requirement, history_id):
    chat_objects = []

    history = None
    if not history_id: 
        history = Chat.objects.create(title=screen_name)
    else:
        try:
            history = Chat.objects.get(id=history_id)
        except FileNotFoundError:
            raise ValueError("History not found.")
    chat = ChatDetail.objects.create(
        chat_id=history,  
        screen_name=screen_name,
        requirement=requirement,
        result='',
        chat_type=2,
        url_requirement='',
        url_result=file_path 
    )

    chat_objects.append(chat)
    return file_name, file_path