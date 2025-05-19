""" 
Module xử lý upload file.
"""

import os
import json
from datetime import datetime, date

import openai
import openpyxl
import logging
import asyncio
import httpx
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from django.conf import settings

from .models import Chat, ChatDetail
from . import views

api_key=""

# Cấu hình logging (ghi log vào file debug.log)
logging.basicConfig(filename="debug.log", level=logging.DEBUG, encoding="utf-8")

@csrf_exempt
def upload_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        # Kiểm tra định dạng file
        if file_extension not in ['xlsx']:
            return JsonResponse({'message': 'Vui lòng chọn file Excel (xlsx)!'}, status=400)
        
        upload_dir = os.path.join(settings.BASE_DIR, 'static', 'upload')

        # Tạo thư mục nếu chưa tồn tại
        os.makedirs(upload_dir, exist_ok=True)

        # Tạo tên file theo thời gian
        original_filename = file.name
        file_path = os.path.join(upload_dir, original_filename)

        # Lưu file vào thư mục `result`
        fs = FileSystemStorage(location=upload_dir)
        filename = fs.get_available_name(original_filename)

        # Lưu file vào thư mục result/
        fs.save(filename, file)
        
        wb = load_workbook(file)

        results = {}
        sheets_data = {}  # Lưu trữ dữ liệu của tất cả các sheet hợp lệ

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            try:
                if validate_file_spec(sheet):
                    data = data_spec(sheet)  # Hàm xử lý dữ liệu Spec
                elif validate_file_api(sheet):
                    data = data_api(sheet)  # Hàm xử lý dữ liệu API
                else:
                    continue  # Nếu sheet không hợp lệ, bỏ qua
                results[sheet_name] = data  # Lưu kết quả của sheet vào dictionary

                if isinstance(data, dict) and "item" in data and data["item"]:
                    sheets_data[sheet_name] = data  # Lưu data hợp lệ vào danh sách

            except Exception as e:
                print(f"⚠️ Lỗi khi xử lý sheet '{sheet_name}': {str(e)}")  
                results[sheet_name] = {"error": str(e)}  # Ghi lỗi vào kết quả của sheet đó

        # Gửi toàn bộ dữ liệu trong 1 lần gọi API
        if sheets_data:
            test_cases = asyncio.run(generate_testcases_async(sheets_data))
            data_test_case = {}
            screen_names = {}
            for sheet_name, result in test_cases.items():
                if "<!DOCTYPE" in result or "<html>" in result:
                    print(f"❌ Lỗi: API trả về HTML thay vì JSON ở sheet {sheet_name}!")
                    continue

                data_test_case[sheet_name] = views.parse_test_cases(result)
                screen_names[sheet_name] = sheets_data[sheet_name].get('screen_name', '')

        history_id = request.POST.get('history_id')
        file_path_requiment = file_path.replace(str(settings.BASE_DIR), "").lstrip("/")
        file_name, file_path_testcase = save_upload(screen_names, data_test_case, history_id, file_path_requiment)

        return JsonResponse({
            "screen_name": next(iter(screen_names.values()), ""),
            "test_cases": test_cases,
            "file_name": file_name,
            "file_path_requiment" : file_path_requiment,
            "file_path_testcase" : file_path_testcase
            },
            status=200)    

    return JsonResponse({'message': 'Không có file nào được tải lên!'}, status=400)

# Hàm tạo prompt cho từng sheet
def build_prompt_for_sheet(sheet_name, data):
    screen_name = data.get('screen_name', '')
    requirement = data.get('requirement', '')
    item = data['item']

    prompt = (
        f"Tôi có màn hình sau, hãy viết ÍT NHẤT 10 test case theo định dạng chuẩn:\n\n"
        f"Sheet: {sheet_name}\n"
        f"Màn hình: {screen_name}\n"
        f"Yêu cầu: {requirement}\n"
        f"Item hiển thị và điều kiện: {item}\n\n"
        "Hãy viết theo format sau:\n"
        "### Test case <số thứ tự>\n"
        "Số thứ tự: <số thứ tự>\n"
        "Độ ưu tiên: <độ ưu tiên của test case>\n"
        "Loại: <loại test case>\n"
        "Mục tiêu: <mục tiêu test case>\n"
        "Dữ liệu kiểm tra: <dữ liệu để test>\n"
        "Điều kiện: <điều kiện để test>\n"
        "Các bước kiểm tra: <các bước để test>\n"
        "Kết quả mong đợi: <kết quả mong đợt>\n"
        "Ghi chú: <ghi chú>\n\n"
        "Ví dụ: Nếu chức năng là 'Đăng nhập', cung cấp các trường hợp kiểm thử như sau:\n\n"
        "### Test case 1\n"
        "Số thứ tự: 1\n"
        "Độ ưu tiên: Cao\n"
        "Loại: Kiểm thử chức năng\n"
        "Mục tiêu: Đăng nhập thành công với thông tin hợp lệ\n"
        "Dữ liệu kiểm tra: Tên người dùng: 'user123', Mật khẩu: 'password123'\n"
        "Điều kiện: Người dùng đã có tài khoản hợp lệ\n"
        "Các bước kiểm tra:\n"
        "    1. Nhập 'user123' vào trường Tên người dùng.\n"
        "    2. Nhập 'password123' vào trường Mật khẩu.\n"
        "    3. Nhấn nút 'Đăng nhập'.\n"
        "Kết quả mong đợi: đăng nhập thành công\n"
        "Ghi chú: Đảm bảo thông tin xác thực hợp lệ.\n"
    )
    return prompt

# Hàm gọi OpenAI API bất đồng bộ dùng httpx
async def call_openai_api_async(prompt, retries=3, delay=3):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    json_data = {
        "model": "gpt-3.5-turbo",
        "messages": [{"role": "user", "content": prompt}],
    }
    for attempt in range(retries):
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(60.0)) as client:
                response = await client.post(url, headers=headers, json=json_data)
                response.raise_for_status()
                data = response.json()
                return data['choices'][0]['message']['content']
        except (httpx.HTTPError, httpx.ReadTimeout) as e:
            logging.debug(f"Attempt {attempt+1} failed: {e}")    
            if attempt < retries - 1:
                await asyncio.sleep(delay)
            else:
                raise e

# Hàm chạy song song nhiều prompt bất đồng bộ
async def generate_testcases_async(data_dict, max_concurrency=5):
    semaphore = asyncio.Semaphore(max_concurrency)
    results = {}

    async def sem_task(sheet_name, data):
        prompt = build_prompt_for_sheet(sheet_name, data)
        async with semaphore:
            try:
                result = await call_openai_api_async(prompt)
            except Exception as e:
                result = f"Error: {str(e)}"
            results[sheet_name] = result

    tasks = [sem_task(sheet_name, data) for sheet_name, data in data_dict.items()]
    await asyncio.gather(*tasks)
    return results

# check validate file spec
def validate_file_spec(sheet):
    expected_screen = "Màn hình"
    expected_screen_requirements = "Yêu cầu màn hình"
    
    screen_cell = sheet["A1"].value
    screen_requirements_cell = sheet["A2"].value
    
    if screen_cell != expected_screen or screen_requirements_cell != expected_screen_requirements:
        return False
    
    expected_headers = ["STT", "Tên Item", "Require", "Type", "Max", "Min", "Condition/Spec/Event", "Data"]
    actual_headers = [sheet.cell(row=8, column=i).value for i in range(1, len(expected_headers) + 1)]
    actual_headers = [header.strip() if header else None for header in actual_headers]
    
    return actual_headers == expected_headers
# check validate file api
def validate_file_api(sheet):
    api_name = "Tên API"
    expected_api_name = sheet["A1"].value
    if api_name != expected_api_name:
        return False
    
    expected_headers = ["Loại", "Tên Tham số", "Bắt buộc", "Mô tả", "Mẫu"]
    actual_headers = []
    column_count = sheet.max_column
    for i in range(1, column_count + 1):
        cell_value = sheet.cell(row=9, column=i).value
        actual_headers.append(cell_value.strip() if cell_value else None)

    # Loại bỏ các phần tử None trong actual_headers
    actual_headers = [header for header in actual_headers if header is not None]

    return actual_headers == expected_headers


# get data trong file excel 
def data_spec(sheet):
    data = {  
        "screen_name": sheet["B1"].value,
        "requirement": sheet["B2"].value,
        "item": []
    }

    # Lấy tiêu đề từ dòng 8 (cột A đến H)
    headers = [cell.value for cell in sheet[8][:8]]

    item = []
    for row in sheet.iter_rows(min_row=9, max_col=8, values_only=True):
        row_data = {}
        for i in range(len(headers)):
            cell_value = row[i] if i < len(row) else None
            if cell_value is not None:
                row_data[headers[i]] = cell_value
        if row_data:  
            item.append(row_data)

    data['item'] = item
    return data  

# get data trong file excel
def data_api(sheet):
    data = {
        "screen_name": sheet["B1"].value,
        "item": []
    }

    # Xác định vị trí của tiêu đề trong sheet
    header_rows = [cell.row for cell in sheet["A"] if cell.value == "Loại"]

    if len(header_rows) < 2:
        return {"error": "Không tìm thấy 'Loại' thứ hai trong cột A"}

    header_index = header_rows[1]  # Lấy dòng chứa 'Loại' thứ 2

    # Lấy tiêu đề cột (dòng 9), chỉ lấy 6 cột đầu tiên
    headers = [cell.value for cell in sheet[9][:6]]
    headers = [header.strip() if header else None for header in headers]  # Chuẩn hóa tiêu đề

    item = []

    for row in sheet.iter_rows(min_row=10, max_row=header_index - 1, max_col=6, values_only=True):
        row_data = {}
        for i, header in enumerate(headers):
            if header:  # Bỏ qua cột không có tiêu đề
                row_data[header] = row[i] if i < len(row) else None

        # Xử lý đặc biệt cho "Tên Tham số"
        if "Tên Tham số" in row_data and any(row_data.values()):
            extra_values = [str(val) for key, val in row_data.items() if key and key != "Tên Tham số" and val]
            if extra_values:
                row_data["Tên Tham số"] = f'{row_data["Tên Tham số"]} gồm {", ".join(extra_values)}'

        if row_data:  # Chỉ thêm dòng có dữ liệu
            item.append(row_data)

    data["item"] = item
    return data


def write_test_case_to_excel(screen_names, test_cases_json):
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_file_path = os.path.join(current_dir, "files", "format-testcase.xlsx")

        if not os.path.exists(template_file_path):
            raise FileNotFoundError("Template file not found!")

        workbook = openpyxl.load_workbook(template_file_path)
        upload_dir = os.path.join(settings.BASE_DIR, 'static', 'result')
        os.makedirs(upload_dir, exist_ok=True)
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{current_time}_testcase.xlsx"
        file_path = os.path.join(upload_dir, file_name)

        # Nếu test_cases_json là chuỗi, parse nó thành dict
        if isinstance(test_cases_json, str):
            try:
                test_cases_json = json.loads(test_cases_json)
            except json.JSONDecodeError as e:
                logging.debug(f"JSON decode error: {e}")
                raise ValueError("Invalid JSON for test_cases_json.")

        # Xác định sheet mẫu (giả sử là sheet đầu tiên)
        template_sheet = workbook.active
        template_sheet_name = template_sheet.title

        # Tạo danh sách các sheet cần giữ lại
        new_sheets = []

        # Duyệt qua từng sheet dựa trên key trong test_cases_json
        for sheet_name, test_cases in test_cases_json.items():
            # Nếu test_cases là chuỗi, parse nó thành list
            if isinstance(test_cases, str):
                try:
                    test_cases = json.loads(test_cases)
                except json.JSONDecodeError as e:
                    logging.debug(f"JSON decode error in sheet {sheet_name}: {e}")
                    raise ValueError(f"Invalid JSON for test_cases in sheet {sheet_name}.")

            # Tạo sheet mới từ sheet mẫu
            sheet = workbook.copy_worksheet(template_sheet)
            sheet.title = sheet_name
            new_sheets.append(sheet_name)  # Lưu lại danh sách sheet hợp lệ

            # Ghi "Tên màn hình" vào ô A2 và ngày hiện tại vào ô F2
            sheet["A2"] = screen_names.get(sheet_name, f"Unnamed_{sheet_name}")
            sheet["F2"] = date.today()

            row = 9
            for case in test_cases:
                # Kiểm tra từng phần tử có phải là dict hay không
                if not isinstance(case, dict):
                    logging.debug(f"Expected dict but got {type(case)} in sheet {sheet_name}: {case}")
                    continue  # Bỏ qua phần tử không hợp lệ

                sheet[f"A{row}"] = case.get("id", "")
                sheet[f"B{row}"] = case.get("priority", "")
                sheet[f"C{row}"] = case.get("type", "")
                sheet[f"D{row}"] = case.get("goal", "")
                # Thay thế dấu phẩy thành dấu phẩy + xuống dòng cho test_data
                sheet[f"E{row}"] = case.get("test_data", "").replace(",", ",\n")
                sheet[f"F{row}"] = case.get("condition", "")
                sheet[f"G{row}"] = case.get("steps", "")
                sheet[f"H{row}"] = case.get("expected_result", "")
                sheet[f"I{row}"] = case.get("note", "")
                row += 1

            # Đặt thuộc tính wrap_text cho các cell từ dòng 9 đến dòng cuối
            for row_cells in sheet.iter_rows(min_row=9, max_row=row - 1):
                for cell in row_cells:
                    cell.alignment = Alignment(wrap_text=True)

        # Xóa sheet mẫu nếu nó vẫn còn trong workbook
        if template_sheet_name in workbook.sheetnames:
            del workbook[template_sheet_name]

        # Lưu file Excel mới
        workbook.save(file_path)
        return file_path, file_name  # Trả về đường dẫn file đã lưu

    except Exception as e:
        print(f"Error: {str(e)}")
        logging.debug(f"Error-1: {str(e)}")
        return None

@csrf_exempt
def save_upload(screen_names, chat_data, history_id, url):
    file_path, file_name = write_test_case_to_excel(screen_names, chat_data)
    file_path = file_path.replace(str(settings.BASE_DIR), "").lstrip("/")
    first_screen_name = next(iter(screen_names.values()), "")
    if not chat_data:
        raise ValueError("No chats provided.")

    chat_objects = []

    history = None
    if not history_id or history_id.lower() == 'null': 
        history = Chat.objects.create(title=first_screen_name)
    else:
        try:
            history = Chat.objects.get(id=history_id)
        except FileNotFoundError:
            raise ValueError("History not found.")
    chat = ChatDetail.objects.create(
        chat_id=history,  
        screen_name=first_screen_name,
        requirement='',
        result='',
        chat_type=2,
        url_requirement=url,
        url_result=file_path 
    )

    chat_objects.append(chat)
    return file_name, file_path
